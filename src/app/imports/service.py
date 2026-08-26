import csv
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from app.core.exceptions import BadRequestException, ForbiddenException
from app.creances.models import StatutCreance
from app.creances.schemas import CreanceCreate
from app.creances.service import CreanceService
from app.debiteurs.repository import DebiteurRepository
from app.debiteurs.schemas import DebiteurCreate
from app.debiteurs.telephone import normaliser as normaliser_telephone
from app.imports.schemas import ImportPreview, ImportResult, ImportRowError
from app.paiements.models import ModePaiement
from app.paiements.schemas import PaiementCreate
from app.paiements.service import PaiementService
from app.users.models import User

# En-tête normalisé -> champ canonique
HEADER_MAP = {
    "nom": "nom",
    "prenom": "prenom",
    "telephone": "telephone",
    "tel": "telephone",
    "phone": "telephone",
    "mobile": "telephone",
    "email": "email",
    "mail": "email",
    "courriel": "email",
    "entreprise": "entreprise",
    "societe": "entreprise",
    "adresse": "adresse",
    "ville": "ville",
    "codepostal": "code_postal",
    "cp": "code_postal",
    # Reference interne du dossier, a distinguer du numero de la facture d'origine.
    "reference": "reference",
    "ref": "reference",
    "numfacture": "numero_facture",
    "numerofacture": "numero_facture",
    "nfacture": "numero_facture",
    "numfact": "numero_facture",
    "facture": "numero_facture",
    "montant": "montant",
    "montantinitial": "montant",
    "montantttc": "montant",
    "montanttotal": "montant",
    "ttc": "montant",
    "montantregle": "montant_regle",
    "montantpaye": "montant_regle",
    "regle": "montant_regle",
    "solde": "montant_restant",
    "resteapayer": "montant_restant",
    "montantrestant": "montant_restant",
    "datefacture": "date_facture",
    "datefact": "date_facture",
    "dateemission": "date_facture",
    "emission": "date_facture",
    "echeance": "echeance",
    "dateecheance": "echeance",
    "statut": "statut",
    "status": "statut",
    "description": "description",
    "libelle": "description",
    # Contexte du dossier — axes de segmentation. « financeur » n'est pas le
    # débiteur : c'est qui doit effectivement payer (famille, bourse, entreprise).
    "etablissement": "etablissement",
    "ecole": "etablissement",
    "structure": "etablissement",
    "site": "etablissement",
    "cycle": "cycle",
    "niveau": "cycle",
    "classe": "cycle",
    "filiere": "cycle",
    "financeur": "financeur",
    "bailleur": "financeur",
    "payeur": "financeur",
    "sourcefinancement": "financeur",
    # Nom du débiteur en un seul champ (ex. société) + personne de contact.
    # « client » et « nomclient » restent acceptés : les fichiers historiques
    # des organisations utilisent encore ce libellé pour désigner le débiteur.
    "debiteur": "nom",
    "nomdebiteur": "nom",
    "client": "nom",
    "raisonsociale": "nom",
    "nomclient": "nom",
    "contact": "prenom",
    "interlocuteur": "prenom",
    "personne": "prenom",
}

# Libellés de statut tolérés (accents/casse/underscore ignorés) -> enum
STATUT_LABELS_MAP = {
    "encours": "EN_COURS",
    "enretard": "EN_RETARD",
    "soldee": "SOLDEE",
    "solde": "SOLDEE",
    "paye": "SOLDEE",
    "payee": "SOLDEE",
    "litige": "LITIGE",
    "annulee": "ANNULEE",
    "annule": "ANNULEE",
}


def _norm_statut(value: object) -> str:
    ascii_ = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z]", "", ascii_.lower())

# Colonnes du modèle Excel téléchargeable (format « factures »)
TEMPLATE_HEADERS = [
    "num_facture",
    "date_facture",
    "debiteur",
    "contact",
    "telephone",
    "email",
    "montant_ttc",
    "montant_regle",
    "date_echeance",
    "statut",
    "description",
    "etablissement",
    "cycle",
    "financeur",
]


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    ascii_ = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_.lower())


def _parse_montant(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        montant = Decimal(str(value))
    else:
        cleaned = re.sub(r"[\s ]", "", str(value)).replace(",", ".")
        montant = Decimal(cleaned)  # peut lever InvalidOperation
    if montant <= 0:
        raise InvalidOperation("montant non positif")
    return montant


def _parse_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("date invalide")


class ImportService:
    def __init__(
        self,
        debiteur_repository: DebiteurRepository,
        creance_service: CreanceService,
        paiement_service: PaiementService,
        current_user: User,
    ) -> None:
        self.debiteur_repository = debiteur_repository
        self.creance_service = creance_service
        self.paiement_service = paiement_service
        self.current_user = current_user

    def _organisation_id(self) -> int:
        if self.current_user.organisation_id is None:
            raise ForbiddenException("Un super-administrateur ne gere pas directement les donnees d'une organisation")
        return self.current_user.organisation_id

    # ---------- Lecture du fichier ----------

    def _parse_file(self, filename: str, content: bytes) -> list[dict]:
        name = (filename or "").lower()
        if name.endswith(".csv"):
            rows = self._read_csv(content)
        elif name.endswith(".xlsx") or name.endswith(".xlsm"):
            rows = self._read_xlsx(content)
        else:
            raise BadRequestException("Format non supporté : utilisez un fichier .xlsx ou .csv")
        if not rows:
            raise BadRequestException("Le fichier est vide ou sans en-tête.")
        return rows

    def _map_headers(self, headers: list) -> list[str]:
        return [HEADER_MAP.get(_norm_header(h), "") for h in headers]

    def _read_xlsx(self, content: bytes) -> list[dict]:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return []
        fields = self._map_headers(list(header_row))
        rows: list[dict] = []
        for excel_index, values in enumerate(it, start=2):
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue
            row = {fields[i]: values[i] for i in range(min(len(fields), len(values))) if fields[i]}
            row["__ligne__"] = excel_index
            rows.append(row)
        return rows

    def _read_csv(self, content: bytes) -> list[dict]:
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:2000]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)
        if not all_rows:
            return []
        fields = self._map_headers(all_rows[0])
        rows: list[dict] = []
        for csv_index, values in enumerate(all_rows[1:], start=2):
            if all(str(v).strip() == "" for v in values):
                continue
            row = {fields[i]: values[i] for i in range(min(len(fields), len(values))) if fields[i]}
            row["__ligne__"] = csv_index
            rows.append(row)
        return rows

    # ---------- Validation ----------

    def _validate_row(self, row: dict) -> tuple[dict | None, str | None]:
        nom = str(row.get("nom") or "").strip()
        if not nom:
            return None, "Nom du débiteur obligatoire"
        prenom = str(row.get("prenom") or "").strip() or "-"

        try:
            montant = _parse_montant(row.get("montant"))
        except (InvalidOperation, ValueError):
            return None, f"Montant invalide : {row.get('montant')!r}"
        if montant is None:
            return None, "Montant obligatoire"

        try:
            echeance = _parse_date(row.get("echeance"))
        except ValueError:
            return None, f"Échéance invalide : {row.get('echeance')!r}"
        if echeance is None:
            return None, "Échéance obligatoire"

        # Date de facture : facultative (les reprises de stock ne l'ont pas toujours),
        # mais si elle est fournie elle doit précéder l'échéance.
        try:
            date_facture = _parse_date(row.get("date_facture"))
        except ValueError:
            return None, f"Date de facture invalide : {row.get('date_facture')!r}"
        if date_facture is not None and date_facture > echeance:
            return None, f"Date de facture ({date_facture}) postérieure à l'échéance ({echeance})"

        # Statut : accepte les libellés FR ("En cours") comme les codes ("EN_COURS").
        statut_input = str(row.get("statut") or "").strip()
        if statut_input:
            code = STATUT_LABELS_MAP.get(_norm_statut(statut_input))
            if code is None and statut_input.upper() in StatutCreance.__members__:
                code = statut_input.upper()
            if code is None:
                return None, f"Statut inconnu : {statut_input}"
            statut = StatutCreance[code]
        else:
            statut = StatutCreance.EN_COURS

        # Montant déjà réglé : colonne "réglé" sinon déduit du solde restant.
        montant_regle = Decimal("0")
        try:
            if row.get("montant_regle") not in (None, ""):
                montant_regle = _parse_montant(row.get("montant_regle")) or Decimal("0")
            elif row.get("montant_restant") not in (None, ""):
                solde = _parse_montant(row.get("montant_restant"))
                if solde is not None:
                    montant_regle = montant - solde
        except (InvalidOperation, ValueError):
            montant_regle = Decimal("0")
        montant_regle = max(Decimal("0"), min(montant_regle, montant))

        telephone = re.sub(r"[\s ]", "", str(row.get("telephone") or "").strip()) or None
        numero_facture = str(row.get("numero_facture") or "").strip() or None
        # À défaut de colonne « référence » dédiée, le numéro de facture sert de
        # référence interne : c'est ce qui fait échouer un ré-import du même fichier
        # sur la contrainte d'unicité, plutôt que de créer des doublons.
        reference = str(row.get("reference") or "").strip() or numero_facture

        record = {
            "debiteur": DebiteurCreate(
                nom=nom,
                prenom=prenom,
                email=str(row.get("email") or "").strip() or None,
                telephone=telephone,
                entreprise=str(row.get("entreprise") or "").strip() or nom,
                adresse=str(row.get("adresse") or "").strip() or None,
                ville=str(row.get("ville") or "").strip() or None,
                code_postal=str(row.get("code_postal") or "").strip() or None,
            ),
            "telephone": telephone,
            "creance": {
                "reference": reference,
                "numero_facture": numero_facture,
                "description": str(row.get("description") or "").strip() or None,
                "montant_initial": montant,
                "date_facture": date_facture,
                "date_echeance": echeance,
                "statut": statut,
                # Colonnes facultatives : absentes du modele standard, elles ne
                # bloquent pas un fichier qui ne les porte pas.
                "etablissement": str(row.get("etablissement") or "").strip() or None,
                "cycle": str(row.get("cycle") or "").strip() or None,
                "financeur": str(row.get("financeur") or "").strip() or None,
            },
            "montant_regle": montant_regle,
        }
        return record, None

    # ---------- Aperçu (dry-run) ----------

    async def preview(self, filename: str, content: bytes) -> ImportPreview:
        rows = self._parse_file(filename, content)
        valides = 0
        erreurs: list[ImportRowError] = []
        for row in rows:
            _, err = self._validate_row(row)
            if err:
                erreurs.append(ImportRowError(ligne=row["__ligne__"], message=err))
            else:
                valides += 1
        return ImportPreview(
            total_lignes=len(rows),
            lignes_valides=valides,
            lignes_invalides=len(erreurs),
            erreurs=erreurs[:200],
        )

    # ---------- Import réel ----------

    async def commit(self, filename: str, content: bytes, dossier_id: int) -> ImportResult:
        """Importe le fichier dans un dossier existant.

        Le dossier est choisi par l'agent, pas devine : un fichier correspond a
        une demande recue d'un client — les trente etudiants d'une ecole entrent
        dans le dossier que cette ecole a confie.
        """
        rows = self._parse_file(filename, content)
        organisation_id = self._organisation_id()
        # 404 si le dossier appartient a une autre organisation.
        await self.creance_service.dossier_service.get_dossier(dossier_id)

        debiteurs_crees = 0
        paiements_repris = 0
        debiteurs_reutilises = 0
        creances_creees = 0
        rejets: list[ImportRowError] = []
        # Caches de la passe en cours, sur la forme canonique du telephone et sur
        # l'email : deux lignes du meme fichier designant la meme personne doivent
        # se rejoindre sans aller-retour en base.
        cache_tel: dict[str, int] = {}
        cache_email: dict[str, int] = {}

        for row in rows:
            ligne = row["__ligne__"]
            record, err = self._validate_row(row)
            if err or record is None:
                rejets.append(ImportRowError(ligne=ligne, message=err or "ligne invalide"))
                continue
            try:
                debiteur_id = None
                telephone = normaliser_telephone(record["telephone"])
                email = (record["debiteur"].email or "").strip().lower() or None

                # Telephone d'abord : c'est l'identifiant le plus souvent renseigne.
                if telephone:
                    if telephone in cache_tel:
                        debiteur_id = cache_tel[telephone]
                        debiteurs_reutilises += 1
                    else:
                        existant = await self.debiteur_repository.get_by_telephone(telephone, organisation_id)
                        if existant is not None:
                            debiteur_id = existant.id
                            debiteurs_reutilises += 1
                            cache_tel[telephone] = debiteur_id

                # Email ensuite : il porte une contrainte d'unicite en base, donc
                # sans ce rattrapage la creation echouerait au lieu de reutiliser.
                if debiteur_id is None and email:
                    if email in cache_email:
                        debiteur_id = cache_email[email]
                        debiteurs_reutilises += 1
                    else:
                        existant = await self.debiteur_repository.get_by_email(email, organisation_id)
                        if existant is not None:
                            debiteur_id = existant.id
                            debiteurs_reutilises += 1
                            cache_email[email] = debiteur_id

                if debiteur_id is None:
                    debiteur = await self.debiteur_repository.create(record["debiteur"], organisation_id)
                    debiteur_id = debiteur.id
                    debiteurs_crees += 1

                if telephone:
                    cache_tel.setdefault(telephone, debiteur_id)
                if email:
                    cache_email.setdefault(email, debiteur_id)

                donnees_creance = dict(record["creance"])
                montant_regle = record.get("montant_regle") or Decimal("0")
                # Une creance qu'on va payer naît EN_COURS, quoi que dise le
                # fichier : c'est le paiement qui la solde, et enregistrer_paiement
                # refuse d'encaisser sur une creance deja SOLDEE. Une ligne
                # « Soldee » avec son reglement complet — le cas le plus normal
                # d'une reprise de stock — etait donc rejetee, en laissant
                # derriere elle une facture marquee soldee dont le restant du
                # valait encore la totalite.
                if montant_regle > 0:
                    donnees_creance["statut"] = StatutCreance.EN_COURS

                creance = await self.creance_service.create_creance(
                    CreanceCreate(debiteur_id=debiteur_id, dossier_id=dossier_id, **donnees_creance)
                )
                creances_creees += 1
                # Montant déjà réglé dans le fichier source : on enregistre un vrai
                # paiement, pas une simple décrémentation du restant. Sans cette ligne
                # dans « paiements », le montant recouvré du tableau de bord ignorerait
                # tout ce qui a été encaissé avant l'entrée du dossier dans l'outil,
                # alors que le restant, lui, en tient compte — les deux chiffres se
                # contrediraient. create_paiement décrémente la créance ET trace
                # l'encaissement, d'où le passage par le service des paiements.
                if montant_regle > 0:
                    await self.paiement_service.create_paiement(
                        PaiementCreate(
                            creance_id=creance.id,
                            montant=montant_regle,
                            # Date de saisie de la créance, c'est-à-dire le jour de
                            # l'import : la date réelle de l'encaissement n'est pas
                            # dans le fichier, on ne l'invente pas.
                            date_paiement=creance.date_saisie,
                            mode_paiement=ModePaiement.REPRISE,
                            notes="Montant déjà réglé, repris du fichier d'import",
                        )
                    )
                    paiements_repris += 1
            except Exception as exc:  # noqa: BLE001 — on rejette la ligne sans casser l'import global
                # Une violation de contrainte laisse la session en echec : sans ce
                # rollback, toutes les lignes suivantes seraient rejetees a leur tour
                # avec « transaction has been rolled back », y compris les valides.
                await self.debiteur_repository.rollback()
                # Le rollback EXPIRE tous les objets ORM de la session, dont
                # current_user. Le premier acces a l'un de ses champs, a la ligne
                # suivante, declenche alors une lecture paresseuse hors contexte
                # async : « greenlet_spawn has not been called ». Toutes les
                # lignes restantes tombaient ainsi, quelle que soit leur validite
                # — une seule ligne fautive faisait perdre tout ce qui la suivait.
                await self.debiteur_repository.db.refresh(self.current_user)
                rejets.append(ImportRowError(ligne=ligne, message=str(exc)[:200]))

        return ImportResult(
            debiteurs_crees=debiteurs_crees,
            debiteurs_reutilises=debiteurs_reutilises,
            creances_creees=creances_creees,
            paiements_repris=paiements_repris,
            lignes_rejetees=rejets,
        )

    # ---------- Modèle Excel ----------

    @staticmethod
    def build_template() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "creances"
        ws.append(TEMPLATE_HEADERS)
        ws.append(
            [
                "FAC-2026-0001",
                "2026-07-31",
                "Atlantique Négoce SA",
                "M. Diallo",
                "+221770000001",
                "compta@atlantique.sn",
                1250000,
                0,
                "2026-08-30",
                "En cours",
                "Facture juillet",
            ]
        )
        ws.append(
            [
                "FAC-2026-0002",
                "2026-08-31",
                "Atlantique Négoce SA",
                "M. Diallo",
                "+221770000001",
                "",
                900000,
                300000,
                "2026-09-30",
                "En cours",
                "Facture août (acompte 300k)",
            ]
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
